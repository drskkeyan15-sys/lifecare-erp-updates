import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
from app_paths import DB_NAME
from icon_loader import get_icon
from money import to_money, split_gst_inclusive, aggregate_gst_by_bill
import theme
import ui_popups

# ReportLab PDF Generation Library
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# openpyxl - for CA/accountant-friendly Excel export (a raw table is
# far more useful to them than a PDF, which they'd have to retype or
# manually re-extract). Same defensive-import pattern as ReportLab
# above - openpyxl is already in the project's tech stack and
# LifeCareERP.spec's hiddenimports, but this guards against a dev
# environment that's missing it, same as the PDF path does.
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class GSTReports:

    def __init__(self, frame):
        self.frame = frame
        self.create_variables()
        self.create_ui()
        self.load_gst_report()

    def create_variables(self):
        self.from_date = tk.StringVar(value=datetime.now().strftime("%Y-%m-01"))
        self.to_date = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.total_taxable = tk.StringVar(value="₹ 0.00")
        self.total_gst = tk.StringVar(value="₹ 0.00")

    def create_ui(self):
        title = tk.Label(
            self.frame,
            text="GST & TAX REPORT SUMMARY (GSTR-1)",
            bg=theme.PRIMARY,
            fg="white",
            font=("Segoe UI", 18, "bold"),
            pady=10
        )
        title.pack(fill="x")

        # ---------------- Filter Frame ----------------
        filter_frame = tk.LabelFrame(
            self.frame,
            text="Date Filter & Export",
            font=("Segoe UI", 10, "bold")
        )
        filter_frame.pack(fill="x", padx=10, pady=10)

        tk.Label(filter_frame, text="From Date (YYYY-MM-DD)").grid(row=0, column=0, padx=5, pady=5)
        tk.Entry(filter_frame, textvariable=self.from_date, width=15).grid(row=0, column=1, padx=5, pady=5)

        tk.Label(filter_frame, text="To Date (YYYY-MM-DD)").grid(row=0, column=2, padx=5, pady=5)
        tk.Entry(filter_frame, textvariable=self.to_date, width=15).grid(row=0, column=3, padx=5, pady=5)

        tk.Button(
            filter_frame,
            text="Generate Report",
            bg="green",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            width=18,
            command=self.load_gst_report,
            cursor="hand2"
        ).grid(row=0, column=4, padx=10, pady=5)

        tk.Button(
            filter_frame,
            text=" Download PDF",
            image=get_icon("download"),
            compound="left",
            bg=theme.ACCENT_PDF_EXPORT,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            padx=12, pady=4,
            command=self.export_pdf,
            cursor="hand2"
        ).grid(row=0, column=5, padx=10, pady=5)

        tk.Button(
            filter_frame,
            text=" Download Excel",
            image=get_icon("download"),
            compound="left",
            bg=theme.STATUS_SUCCESS,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            padx=12, pady=4,
            command=self.export_excel,
            cursor="hand2"
        ).grid(row=0, column=6, padx=10, pady=5)

        # ---------------- Summary Frame ----------------
        summary_frame = tk.Frame(self.frame)
        summary_frame.pack(fill="x", padx=10, pady=10)

        tk.Label(summary_frame, text="Total Taxable Value:", font=("Segoe UI", 11, "bold")).pack(side="left", padx=5)
        tk.Label(summary_frame, textvariable=self.total_taxable, fg="blue", font=("Segoe UI", 11, "bold")).pack(side="left", padx=10)

        tk.Label(summary_frame, text="Total GST Collected:", font=("Segoe UI", 11, "bold")).pack(side="left", padx=20)
        tk.Label(summary_frame, textvariable=self.total_gst, fg="green", font=("Segoe UI", 11, "bold")).pack(side="left", padx=10)

        # ---------------- Table Frame ----------------
        table_frame = tk.Frame(self.frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)

        columns = ("Bill No", "Date", "Customer", "Taxable Amount", "CGST", "SGST", "Total Tax")
        self.gstTable = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            height=15,
            style="ERP.Treeview"
        )

        for c in columns:
            self.gstTable.heading(c, text=c)
            self.gstTable.column(c, width=130, anchor="center")

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.gstTable.yview)
        self.gstTable.configure(yscrollcommand=scrollbar.set)
        self.gstTable.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def load_gst_report(self):
        f_date = self.from_date.get().strip()
        t_date = self.to_date.get().strip()

        self.gstTable.delete(*self.gstTable.get_children())
        con = sqlite3.connect(DB_NAME)
        cur = con.cursor()

        try:
            cur.execute("""
                SELECT bill_no, bill_date, customer, subtotal, total
                FROM sales
                WHERE bill_date BETWEEN ? AND ?
            """, (f_date, t_date))
            rows = cur.fetchall()

            # BUG FIX: this used to do `subtotal / 1.12` for every single
            # bill - i.e. it assumed every medicine in every bill was
            # taxed at a flat 12% GST. Real pharmacy stock spans the 5%,
            # 12% and 18% slabs too, so any bill containing a 5%- or
            # 18%-rated medicine got a wrong taxable/GST split here (and
            # therefore a wrong number on this compliance report).
            # Now each sold line is split at ITS OWN medicine's actual
            # medicine_master.gst rate (joined by name, since sales_items
            # only stores the medicine name, not its id - see
            # SOFTWARE_OVERVIEW.md) and the per-line splits are summed
            # per bill. A medicine that can't be matched (renamed/deleted
            # since the sale) falls back to the old flat-12% estimate for
            # just that one line, so a report row never silently drops to
            # ₹0 - it degrades to the previous behaviour instead.
            cur.execute("""
                SELECT si.bill_no, si.total, mm.gst
                FROM sales_items si
                JOIN sales s ON s.bill_no = si.bill_no
                LEFT JOIN medicine_master mm ON mm.name = si.medicine
                WHERE s.bill_date BETWEEN ? AND ?
            """, (f_date, t_date))
            gst_by_bill = aggregate_gst_by_bill(cur.fetchall(), fallback_rate=12)

            tot_taxable = 0.0
            tot_gst_val = 0.0

            for r in rows:
                bill_no, bill_date, customer, subtotal, total = r
                cust_name = customer if customer else "Walk-in Customer"
                sub = float(subtotal or 0.0)

                taxable, gst_amount = gst_by_bill.get(bill_no, (None, None))
                if taxable is None:
                    # No matching sales_items rows found for this bill at
                    # all (e.g. very old data) - fall back to the
                    # original flat-12% estimate rather than showing 0.
                    taxable, gst_amount = split_gst_inclusive(sub, 12)
                cgst = to_money(gst_amount / 2)
                sgst = to_money(gst_amount - cgst)

                tot_taxable += taxable
                tot_gst_val += gst_amount

                self.gstTable.insert("", "end", values=(
                    bill_no, bill_date, cust_name,
                    f"₹ {taxable:.2f}", f"₹ {cgst:.2f}", f"₹ {sgst:.2f}", f"₹ {gst_amount:.2f}"
                ))

            tot_taxable = to_money(tot_taxable)
            tot_gst_val = to_money(tot_gst_val)
            self.total_taxable.set(f"₹ {tot_taxable:,.2f}")
            self.total_gst.set(f"₹ {tot_gst_val:,.2f}")

        except Exception as e:
            ui_popups.show_error(self.frame, "Error", str(e))
        finally:
            con.close()

    def export_pdf(self):
        if not REPORTLAB_AVAILABLE:
            ui_popups.show_error(self.frame, "Missing Library", "ReportLab library is not installed. Run 'pip install reportlab' in terminal.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=f"GST_Report_{self.from_date.get()}_to_{self.to_date.get()}.pdf"
        )

        if not file_path:
            return

        try:
            c = canvas.Canvas(file_path, pagesize=letter)
            width, height = letter

            # Title
            c.setFont("Helvetica-Bold", 16)
            c.drawString(50, height - 50, "Life Care Pharma V2 - GST Report")
            
            c.setFont("Helvetica", 10)
            c.drawString(50, height - 70, f"Period: {self.from_date.get()} to {self.to_date.get()}")
            c.drawString(50, height - 85, f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

            # Summary
            c.setFont("Helvetica-Bold", 11)
            c.drawString(50, height - 110, f"Total Taxable Value: {self.total_taxable.get()}")
            c.drawString(250, height - 110, f"Total GST Collected: {self.total_gst.get()}")

            # Table Header
            y = height - 140
            c.setFont("Helvetica-Bold", 9)
            c.drawString(50, y, "Bill No")
            c.drawString(140, y, "Date")
            c.drawString(220, y, "Customer")
            c.drawString(330, y, "Taxable")
            c.drawString(410, y, "CGST")
            c.drawString(470, y, "SGST")
            c.drawString(530, y, "Total Tax")

            c.line(50, y - 5, 570, y - 5)
            y -= 20

            c.setFont("Helvetica", 9)
            for row_id in self.gstTable.get_children():
                vals = self.gstTable.item(row_id)["values"]
                if y < 50:
                    c.showPage()
                    y = height - 50
                
                c.drawString(50, y, str(vals[0]))
                c.drawString(140, y, str(vals[1]))
                c.drawString(220, y, str(vals[2][:15]))
                c.drawString(330, y, str(vals[3]))
                c.drawString(410, y, str(vals[4]))
                c.drawString(470, y, str(vals[5]))
                c.drawString(530, y, str(vals[6]))
                y -= 18

            c.save()
            ui_popups.show_info(self.frame, "Success", f"GST Report PDF saved successfully at:\n{file_path}")
        except Exception as e:
            ui_popups.show_error(self.frame, "PDF Error", str(e))

    def export_excel(self):
        """
        Same data as the on-screen table/PDF, as a real .xlsx - meant
        for handing to a CA/accountant, who can filter/pivot/re-total
        it directly instead of retyping numbers off a PDF.
        """
        if not OPENPYXL_AVAILABLE:
            ui_popups.show_error(self.frame, "Missing Library", "openpyxl library is not installed. Run 'pip install openpyxl' in terminal.")
            return

        if not self.gstTable.get_children():
            ui_popups.show_warning(self.frame, "No Data", "Generate the report first (nothing to export).")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"GST_Report_{self.from_date.get()}_to_{self.to_date.get()}.xlsx"
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "GST Report"

            ws["A1"] = "Life Care Pharma V2 - GST Report"
            ws["A1"].font = Font(size=14, bold=True)
            ws["A2"] = f"Period: {self.from_date.get()} to {self.to_date.get()}"
            ws["A3"] = f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws["A4"] = f"Total Taxable Value: {self.total_taxable.get()}"
            ws["A5"] = f"Total GST Collected: {self.total_gst.get()}"

            headers = ("Bill No", "Date", "Customer", "Taxable Amount", "CGST", "SGST", "Total Tax")
            header_row = 7
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            row_num = header_row + 1
            for row_id in self.gstTable.get_children():
                vals = self.gstTable.item(row_id)["values"]
                for col, val in enumerate(vals, start=1):
                    ws.cell(row=row_num, column=col, value=val)
                row_num += 1

            # Reasonable default column widths - openpyxl doesn't
            # auto-size, and the default 8.43 truncates every column here.
            widths = (14, 12, 22, 16, 12, 12, 12)
            for col, width in enumerate(widths, start=1):
                ws.column_dimensions[chr(64 + col)].width = width

            wb.save(file_path)
            ui_popups.show_info(self.frame, "Success", f"GST Report Excel saved successfully at:\n{file_path}")
        except Exception as e:
            ui_popups.show_error(self.frame, "Excel Error", str(e))