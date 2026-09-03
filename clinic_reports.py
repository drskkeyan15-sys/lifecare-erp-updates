import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime

import clinic_repository as repo
import theme

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import csv
import ui_popups


def _summary_rows(summary):
    """Turns a range_summary()-shaped dict into (Metric, Value) rows so
    the Daily/Monthly/Yearly reports can reuse the exact same generic
    Treeview + export machinery every other report here uses.

    Three profit rows on purpose (official names, see clinic_repository.
    compute_profit_breakdown()'s docstring for full definitions):
        Consulting Charge      = Total Collection - Medicine MRP Value
        Actual Net Profit      = Total Collection - Purchase Cost (the
                                  real money made)
        Medicine Margin Profit = Medicine MRP Value - Purchase Cost (what
                                  the margin would be at full printed MRP)
    Net Profit (last row) is a separate, further figure: Actual Net
    Profit minus Operating Expenses for the period."""
    return [
        ("Patients (Visits)", summary["visits"]),
        ("Unique Patients", summary["unique_patients"]),
        ("Consultation Income", f"{summary['consultation_income']:.2f}"),
        ("Medicine MRP Value", f"{summary['medicine_mrp_value']:.2f}"),
        ("Medicine Purchase Cost", f"{summary['medicine_purchase_cost']:.2f}"),
        ("Total Collection", f"{summary['total_collection']:.2f}"),
        ("Consulting Charge", f"{summary['consulting_charge']:.2f}"),
        ("Actual Net Profit", f"{summary['actual_net_profit']:.2f}"),
        ("Medicine Margin Profit", f"{summary['medicine_margin_profit']:.2f}"),
        ("Expenses", f"{summary['expenses']:.2f}"),
        ("Net Profit (after Expenses)", f"{summary['net_profit']:.2f}"),
    ]


# Each report: display name -> (columns, loader(date_from, date_to, patient_id) -> rows)
# `date_from`/`date_to` are plain 'YYYY-MM-DD' from the two entry boxes;
# each loader converts to the '...00:00:00'/'...23:59:59' shape
# clinic_repository's date-range functions expect.
def _daily(date_from, date_to, patient_id):
    return _summary_rows(repo.daily_report(date_from))


def _monthly(date_from, date_to, patient_id):
    d = datetime.strptime(date_from, "%Y-%m-%d")
    summary = repo.monthly_report(d.year, d.month)
    rows = _summary_rows(summary)
    rows.append(("Avg Collection / Patient", f"{summary['avg_collection_per_patient']:.2f}"))
    rows.append(("Avg Profit / Patient", f"{summary['avg_profit_per_patient']:.2f}"))
    rows.append(("Highest Collection Day", str(summary["highest_day"])))
    rows.append(("Lowest Collection Day", str(summary["lowest_day"])))
    return rows


def _yearly(date_from, date_to, patient_id):
    d = datetime.strptime(date_from, "%Y-%m-%d")
    summary = repo.yearly_report(d.year)
    rows = _summary_rows(summary)
    for i, m in enumerate(summary["monthly_breakdown"], start=1):
        rows.append((f"Month {i:02d} Collection", f"{m['total_collection']:.2f}"))
        rows.append((f"Month {i:02d} Net Profit", f"{m['net_profit']:.2f}"))
    return rows


def _dt_range(date_from, date_to):
    return f"{date_from} 00:00:00", f"{date_to} 23:59:59"


REPORTS = {
    "Daily Clinic Report": (("Metric", "Value"), _daily),
    "Monthly Clinic Report": (("Metric", "Value"), _monthly),
    "Yearly Clinic Report": (("Metric", "Value"), _yearly),
    "Patient Visit Report": (
        ("Visit No", "Patient", "Doctor", "Date", "Consultation", "Collection", "Actual Net Profit", "Status"),
        lambda f, t, p: repo.patient_visit_report(*_dt_range(f, t))
    ),
    "Medicine Usage Report": (
        ("Medicine", "Qty Used", "Purchase Cost", "MRP Value", "Medicine Margin Profit"),
        lambda f, t, p: repo.medicine_usage_report(*_dt_range(f, t))
    ),
    "Injection Usage Report": (
        ("Injection", "Qty Used", "Purchase Cost", "MRP Value", "Medicine Margin Profit"),
        lambda f, t, p: repo.injection_usage_report(*_dt_range(f, t))
    ),
    "Medicine Cost Report": (
        ("Item", "Type", "Qty Used", "Purchase Cost"),
        lambda f, t, p: repo.medicine_cost_report(*_dt_range(f, t))
    ),
    "Profit Breakdown Report": (
        ("Visit No", "Date", "Patient", "Consultation", "MRP Value", "Total Collection",
         "Purchase Cost", "Actual Net Profit", "Consulting Charge", "Medicine Margin Profit"),
        lambda f, t, p: repo.gross_profit_report(*_dt_range(f, t))
    ),
    "Net Profit Report": (
        ("Date", "Revenue", "Actual Net Profit", "Expenses", "Net Profit"),
        lambda f, t, p: repo.net_profit_report(*_dt_range(f, t))
    ),
    "Expense Report": (
        ("ID", "Date", "Category", "Description", "Amount", "Payment Mode"),
        lambda f, t, p: repo.expense_report(f, t)[0]
    ),
    "Doctor-wise Report": (
        ("Doctor", "Visits", "Consultation", "Actual Net Profit", "Total Collection"),
        lambda f, t, p: repo.doctor_report(*_dt_range(f, t))
    ),
    "Patient-wise Treatment History": (
        ("Visit ID", "Visit No", "Date", "Doctor", "Consultation", "Collection", "Actual Net Profit", "Status"),
        lambda f, t, p: repo.patient_history_report(p)
    ),
    "Stock Used in Clinic Report": (
        ("Item", "Batch", "Type", "Qty Used"),
        lambda f, t, p: repo.stock_used_report(*_dt_range(f, t))
    ),
}


class ClinicReports:
    """13 Clinic Ledger reports (CLINIC_LEDGER_WORKFLOW.md's reports
    section) behind one screen - a single generic Treeview + Export
    PDF/Excel/CSV trio driven by whatever columns/rows the selected
    report loaded, same "export straight from the Treeview" trick
    gst_reports.py's export_pdf()/export_excel() already use, just made
    column-agnostic here since 13 different report shapes would
    otherwise need 13 near-duplicate export functions."""

    def __init__(self, frame, on_close=None):
        self.frame = frame
        self.on_close = on_close
        self.create_variables()
        self.create_ui()

    def create_variables(self):
        today = datetime.now().strftime("%Y-%m-%d")
        self.report_name = tk.StringVar(value="Daily Clinic Report")
        self.date_from = tk.StringVar(value=today[:8] + "01")
        self.date_to = tk.StringVar(value=today)
        self.patient_id = tk.StringVar()

    def create_ui(self):
        title = tk.Label(
            self.frame, text="CLINIC REPORTS",
            bg=theme.PRIMARY, fg="white", font=("Segoe UI", 18, "bold"), pady=10
        )
        title.pack(fill="x")

        filt = tk.LabelFrame(self.frame, text="Report Filter", font=("Segoe UI", 10, "bold"))
        filt.pack(fill="x", padx=10, pady=10)

        tk.Label(filt, text="Report").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Combobox(filt, textvariable=self.report_name, values=list(REPORTS.keys()),
                     state="readonly", width=32).grid(row=0, column=1, padx=5)

        tk.Label(filt, text="From").grid(row=0, column=2, padx=5)
        tk.Entry(filt, textvariable=self.date_from, width=12).grid(row=0, column=3)
        tk.Label(filt, text="To").grid(row=0, column=4, padx=5)
        tk.Entry(filt, textvariable=self.date_to, width=12).grid(row=0, column=5)

        tk.Label(filt, text="Patient ID (Patient History only)").grid(row=0, column=6, padx=5)
        tk.Entry(filt, textvariable=self.patient_id, width=8).grid(row=0, column=7)

        tk.Button(filt, text="Generate", bg=theme.PRIMARY, fg="white",
                  command=self.generate).grid(row=0, column=8, padx=10)

        exp = tk.Frame(self.frame)
        exp.pack(fill="x", padx=10)
        tk.Button(exp, text="Export CSV", bg=theme.ACCENT_NEUTRAL, fg="white",
                  command=self.export_csv).pack(side="left", padx=5)
        tk.Button(exp, text="Export Excel", bg=theme.STATUS_SUCCESS, fg="white",
                  command=self.export_excel).pack(side="left", padx=5)
        tk.Button(exp, text="Export PDF", bg=theme.ACCENT_PDF_EXPORT, fg="white",
                  command=self.export_pdf).pack(side="left", padx=5)
        if self.on_close:
            tk.Button(exp, text="Close", bg=theme.STATUS_DANGER, fg="white",
                      command=self.on_close).pack(side="right", padx=5)

        table = tk.Frame(self.frame)
        table.pack(fill="both", expand=True, padx=10, pady=10)
        self.reportTable = ttk.Treeview(table, show="headings", height=18, style="ERP.Treeview")
        self.reportTable.pack(fill="both", expand=True)

    def generate(self):
        name = self.report_name.get()
        columns, loader = REPORTS[name]
        try:
            patient_id = int(self.patient_id.get()) if self.patient_id.get().strip() else None
            if name == "Patient-wise Treatment History" and not patient_id:
                ui_popups.show_error(self.frame, "Error", "Enter a Patient ID for this report")
                return
            rows = loader(self.date_from.get().strip(), self.date_to.get().strip(), patient_id)
        except Exception as e:
            ui_popups.show_error(self.frame, "Error", str(e))
            return

        self.reportTable.delete(*self.reportTable.get_children())
        self.reportTable["columns"] = columns
        for c in columns:
            self.reportTable.heading(c, text=c)
            self.reportTable.column(c, width=max(90, 700 // len(columns)), anchor="center")
        for row in rows:
            self.reportTable.insert("", "end", values=row)

        if not rows:
            ui_popups.show_info(self.frame, "No Data", "No records found for this filter.")

    def _current_columns_and_rows(self):
        columns = self.reportTable["columns"]
        rows = [self.reportTable.item(i)["values"] for i in self.reportTable.get_children()]
        return columns, rows

    def export_csv(self):
        columns, rows = self._current_columns_and_rows()
        if not rows:
            ui_popups.show_warning(self.frame, "No Data", "Generate the report first (nothing to export).")
            return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV files", "*.csv")],
            initialfile=f"{self.report_name.get().replace(' ', '_')}.csv"
        )
        if not file_path:
            return
        try:
            with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                writer.writerows(rows)
            ui_popups.show_info(self.frame, "Success", f"CSV saved at:\n{file_path}")
        except Exception as e:
            ui_popups.show_error(self.frame, "CSV Error", str(e))

    def export_excel(self):
        if not OPENPYXL_AVAILABLE:
            ui_popups.show_error(self.frame, "Missing Library", "openpyxl library is not installed. Run 'pip install openpyxl' in terminal.")
            return
        columns, rows = self._current_columns_and_rows()
        if not rows:
            ui_popups.show_warning(self.frame, "No Data", "Generate the report first (nothing to export).")
            return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{self.report_name.get().replace(' ', '_')}.xlsx"
        )
        if not file_path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = self.report_name.get()[:31]
            ws["A1"] = f"Life Care Pharmacy - {self.report_name.get()}"
            ws["A1"].font = Font(size=14, bold=True)
            ws["A2"] = f"Period: {self.date_from.get()} to {self.date_to.get()}"
            ws["A3"] = f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            header_row = 5
            for col, header in enumerate(columns, start=1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            row_num = header_row + 1
            for row in rows:
                for col, val in enumerate(row, start=1):
                    ws.cell(row=row_num, column=col, value=val)
                row_num += 1
            for col in range(1, len(columns) + 1):
                ws.column_dimensions[chr(64 + col)].width = 18
            wb.save(file_path)
            ui_popups.show_info(self.frame, "Success", f"Excel saved at:\n{file_path}")
        except Exception as e:
            ui_popups.show_error(self.frame, "Excel Error", str(e))

    def export_pdf(self):
        if not REPORTLAB_AVAILABLE:
            ui_popups.show_error(self.frame, "Missing Library", "ReportLab library is not installed. Run 'pip install reportlab' in terminal.")
            return
        columns, rows = self._current_columns_and_rows()
        if not rows:
            ui_popups.show_warning(self.frame, "No Data", "Generate the report first (nothing to export).")
            return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")],
            initialfile=f"{self.report_name.get().replace(' ', '_')}.pdf"
        )
        if not file_path:
            return
        try:
            c = canvas.Canvas(file_path, pagesize=letter)
            width, height = letter
            c.setFont("Helvetica-Bold", 14)
            c.drawString(50, height - 50, f"Life Care Pharmacy - {self.report_name.get()}")
            c.setFont("Helvetica", 9)
            c.drawString(50, height - 68, f"Period: {self.date_from.get()} to {self.date_to.get()}")

            col_width = min(90, (width - 100) / max(1, len(columns)))
            y = height - 100
            c.setFont("Helvetica-Bold", 8)
            for i, col in enumerate(columns):
                c.drawString(50 + i * col_width, y, str(col)[:14])
            c.line(50, y - 5, width - 50, y - 5)
            y -= 18

            c.setFont("Helvetica", 8)
            for row in rows:
                if y < 50:
                    c.showPage()
                    y = height - 50
                for i, val in enumerate(row):
                    c.drawString(50 + i * col_width, y, str(val)[:14])
                y -= 15
            c.save()
            ui_popups.show_info(self.frame, "Success", f"PDF saved at:\n{file_path}")
        except Exception as e:
            ui_popups.show_error(self.frame, "PDF Error", str(e))
